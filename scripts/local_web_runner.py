#!/usr/bin/env python3
"""
Local Development Command Center

Unified web-based UI for ARRMS integration local development.
Provides:
- Embedded mock Onspring server
- Direct execution of actual handlers (src/handlers/*)
- Real-time log streaming
- Configuration management
- Mock data viewing/editing

Usage:
    python scripts/local_web_runner.py
    python scripts/local_web_runner.py --port 8080 --env-file .env.local

Then open http://localhost:8080 in your browser.
"""

import argparse
import io
import json
import logging
import os
import queue
import sys
import tempfile
import threading
import time
import traceback
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from typing import Any, Dict, List, Optional
from urllib.parse import parse_qs, urlparse

# Add stubs directory FIRST to override aws_lambda_powertools with local stubs
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "stubs"))
# Add scripts directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
# Add src directory to path for handler imports
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))

from local_runner import LocalARRMSClient, transform_record, load_env_file
from mock_onspring import MockOnspringClient, MOCK_RECORDS, MOCK_COMPANIES

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("command_center")

# Suppress noisy loggers
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("werkzeug").setLevel(logging.WARNING)


# ============================================================================
# Log Buffer for Real-time Streaming
# ============================================================================


class LogBuffer:
    """Thread-safe buffer for log messages that can be streamed via SSE."""

    def __init__(self, max_size: int = 1000):
        self.messages: List[Dict[str, Any]] = []
        self.max_size = max_size
        self.lock = threading.Lock()
        self.subscribers: List[queue.Queue] = []

    def add(self, level: str, message: str, source: str = "SYSTEM", data: Optional[Dict] = None):
        """Add a log message and notify subscribers."""
        entry = {
            "timestamp": datetime.now().isoformat(),
            "level": level,
            "source": source,
            "message": message,
            "data": data,
        }

        with self.lock:
            self.messages.append(entry)
            if len(self.messages) > self.max_size:
                self.messages = self.messages[-self.max_size :]

            # Notify all SSE subscribers
            for q in self.subscribers:
                try:
                    q.put_nowait(entry)
                except queue.Full:
                    pass

    def subscribe(self) -> queue.Queue:
        """Create a new subscriber queue for SSE."""
        q = queue.Queue(maxsize=100)
        with self.lock:
            self.subscribers.append(q)
        return q

    def unsubscribe(self, q: queue.Queue):
        """Remove a subscriber queue."""
        with self.lock:
            if q in self.subscribers:
                self.subscribers.remove(q)

    def get_recent(self, count: int = 50) -> List[Dict]:
        """Get recent log messages."""
        with self.lock:
            return self.messages[-count:]

    def clear(self):
        """Clear all log messages."""
        with self.lock:
            self.messages.clear()

    # Convenience methods
    def info(self, message: str, source: str = "SYSTEM", data: Optional[Dict] = None):
        self.add("INFO", message, source, data)
        logger.info(f"[{source}] {message}")

    def error(self, message: str, source: str = "SYSTEM", data: Optional[Dict] = None):
        self.add("ERROR", message, source, data)
        logger.error(f"[{source}] {message}")

    def warning(self, message: str, source: str = "SYSTEM", data: Optional[Dict] = None):
        self.add("WARNING", message, source, data)
        logger.warning(f"[{source}] {message}")

    def debug(self, message: str, source: str = "SYSTEM", data: Optional[Dict] = None):
        self.add("DEBUG", message, source, data)
        logger.debug(f"[{source}] {message}")

    def handler(self, message: str, data: Optional[Dict] = None):
        """Log from handler execution."""
        self.add("HANDLER", message, "HANDLER", data)
        logger.info(f"[HANDLER] {message}")

    def onspring(self, message: str, data: Optional[Dict] = None):
        """Log from mock Onspring."""
        self.add("ONSPRING", message, "ONSPRING", data)
        logger.info(f"[ONSPRING] {message}")

    def arrms(self, message: str, data: Optional[Dict] = None):
        """Log from ARRMS client."""
        self.add("ARRMS", message, "ARRMS", data)
        logger.info(f"[ARRMS] {message}")

    def webhook(self, message: str, data: Optional[Dict] = None):
        """Log webhook events."""
        self.add("WEBHOOK", message, "WEBHOOK", data)
        logger.info(f"[WEBHOOK] {message}")


# Global log buffer
log_buffer = LogBuffer()


# ============================================================================
# Embedded Mock Onspring Server
# ============================================================================


class MockOnspringServer:
    """Embedded mock Onspring server running in a background thread."""

    def __init__(self, port: int = 5001, host: str = "127.0.0.1"):
        self.port = port
        self.host = host
        self.server = None
        self.thread = None
        self.running = False
        self.field_updates: List[Dict[str, Any]] = []

    def start(self):
        """Start the mock server in a background thread."""
        if self.running:
            return

        try:
            from flask import Flask, jsonify, request, send_file

            app = Flask(__name__)
            app.logger.setLevel(logging.WARNING)

            # Store reference to self for routes
            mock_server = self

            @app.route("/Ping", methods=["GET"])
            def ping():
                return jsonify({"status": "ok", "message": "Mock Onspring Server"})

            @app.route("/Records/appId/<int:app_id>/recordId/<int:record_id>", methods=["GET"])
            def get_record(app_id: int, record_id: int):
                log_buffer.onspring(f"GET record {record_id} from app {app_id}")

                # Check companies (app 249)
                if app_id == 249 and record_id in MOCK_COMPANIES:
                    return jsonify(MOCK_COMPANIES[record_id])

                # Check questionnaire records
                if record_id in MOCK_RECORDS:
                    return jsonify(MOCK_RECORDS[record_id])

                # Return first record if not found
                if MOCK_RECORDS:
                    log_buffer.onspring(f"Record {record_id} not found, returning first mock")
                    return jsonify(list(MOCK_RECORDS.values())[0])

                return jsonify({"error": "Record not found"}), 404

            @app.route("/Records", methods=["PUT"])
            def update_record():
                data = request.get_json()
                app_id = data.get("appId")
                record_id = data.get("recordId")
                fields = data.get("fields", {})

                log_buffer.onspring(f"PUT record {record_id} - updating {len(fields)} fields", {"fields": fields})

                # Track update
                mock_server.field_updates.append({
                    "timestamp": datetime.now().isoformat(),
                    "app_id": app_id,
                    "record_id": record_id,
                    "fields": fields,
                })

                # Update mock record
                if record_id in MOCK_RECORDS:
                    for field_id_str, value in fields.items():
                        field_id = int(field_id_str)
                        for field in MOCK_RECORDS[record_id].get("fieldData", []):
                            if field.get("fieldId") == field_id:
                                field["value"] = value
                                break

                return jsonify({"id": record_id, "warnings": []})

            @app.route("/Files/recordId/<int:record_id>/fieldId/<int:field_id>/fileId/<int:file_id>/file", methods=["GET"])
            def download_file(record_id: int, field_id: int, file_id: int):
                log_buffer.onspring(f"GET file {file_id} from record {record_id}")

                # Generate mock Excel content
                content = mock_server._generate_mock_excel()
                return send_file(
                    io.BytesIO(content),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name="questionnaire.xlsx",
                )

            @app.route("/Files/recordId/<int:record_id>/fieldId/<int:field_id>/fileId/<int:file_id>", methods=["GET"])
            def get_file_info(record_id: int, field_id: int, file_id: int):
                return jsonify({
                    "fileId": file_id,
                    "fileName": f"questionnaire_{file_id}.xlsx",
                    "fileSize": 15000,
                    "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                })

            def run_server():
                from werkzeug.serving import make_server

                self.server = make_server(self.host, self.port, app, threaded=True)
                self.running = True
                log_buffer.info(f"Mock Onspring server started on http://{self.host}:{self.port}", "ONSPRING")
                self.server.serve_forever()

            self.thread = threading.Thread(target=run_server, daemon=True)
            self.thread.start()

            # Wait for server to start
            time.sleep(0.5)

        except ImportError:
            log_buffer.error("Flask not installed - mock Onspring server disabled", "ONSPRING")
            self.running = False

    def stop(self):
        """Stop the mock server."""
        if self.server:
            self.server.shutdown()
            self.running = False
            log_buffer.info("Mock Onspring server stopped", "ONSPRING")

    def _generate_mock_excel(self) -> bytes:
        """Generate mock Excel file content."""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Questionnaire"
            ws["A1"] = "Question"
            ws["B1"] = "Response"
            ws["A2"] = "What is your company name?"
            ws["A3"] = "Describe your security policies"
            ws["A4"] = "Do you have SOC 2 certification?"

            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
        except ImportError:
            return b"Mock Excel Content"

    def get_updates(self) -> List[Dict]:
        """Get list of field updates made to mock Onspring."""
        return self.field_updates

    def clear_updates(self):
        """Clear the update history."""
        self.field_updates.clear()


# ============================================================================
# Handler Execution
# ============================================================================


class HandlerExecutor:
    """Executes actual handler code with local configuration."""

    def __init__(self, mock_onspring_url: str, arrms_client: LocalARRMSClient):
        self.mock_onspring_url = mock_onspring_url
        self.arrms_client = arrms_client

    def create_mock_context(self):
        """Create a mock Lambda context."""

        class MockContext:
            function_name = "local-dev"
            function_version = "$LATEST"
            invoked_function_arn = "arn:aws:lambda:local:000000000000:function:local-dev"
            memory_limit_in_mb = 512
            aws_request_id = f"local-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            log_group_name = "/local/dev"
            log_stream_name = "local"

            @staticmethod
            def get_remaining_time_in_millis():
                return 300000

        return MockContext()

    def run_webhook_handler(self, record_id: int, app_id: int) -> Dict[str, Any]:
        """Run the actual webhook handler code."""
        log_buffer.handler(f"Executing webhook handler for record {record_id}")

        try:
            # Set up environment for handler
            os.environ["LOCAL_DEV"] = "true"
            os.environ["ONSPRING_API_URL"] = self.mock_onspring_url
            os.environ["ONSPRING_API_KEY_SECRET"] = "onspring-api-key"
            os.environ["ONSPRING_API_KEY"] = "local-test-key"
            os.environ["ARRMS_API_KEY_SECRET"] = "arrms-api-key"
            os.environ.setdefault("ONSPRING_DEFAULT_APP_ID", str(app_id))

            # Create webhook event
            event = {
                "httpMethod": "POST",
                "path": "/webhook/onspring",
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps([{"RecordId": str(record_id), "AppId": str(app_id)}]),
                "isBase64Encoded": False,
            }

            # Import and run handler
            from handlers.onspring_webhook import lambda_handler

            result = lambda_handler(event, self.create_mock_context())

            log_buffer.handler(f"Webhook handler completed", {"status_code": result.get("statusCode")})
            return result

        except Exception as e:
            tb = traceback.format_exc()
            log_buffer.error(f"Handler execution failed: {str(e)}", "HANDLER", {"traceback": tb})
            return {"statusCode": 500, "body": json.dumps({"error": str(e)})}

    def run_sync_handler(self, app_id: int) -> Dict[str, Any]:
        """Run the actual sync handler code."""
        log_buffer.handler(f"Executing sync handler for app {app_id}")

        try:
            os.environ["LOCAL_DEV"] = "true"
            os.environ["ONSPRING_API_URL"] = self.mock_onspring_url
            os.environ["ONSPRING_API_KEY_SECRET"] = "onspring-api-key"
            os.environ["ONSPRING_API_KEY"] = "local-test-key"
            os.environ["ARRMS_API_KEY_SECRET"] = "arrms-api-key"

            event = {
                "httpMethod": "POST",
                "path": "/sync/onspring-to-arrms",
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"app_id": app_id, "batch_size": 10}),
                "isBase64Encoded": False,
            }

            from handlers.onspring_to_arrms import lambda_handler

            result = lambda_handler(event, self.create_mock_context())

            log_buffer.handler(f"Sync handler completed", {"status_code": result.get("statusCode")})
            return result

        except Exception as e:
            tb = traceback.format_exc()
            log_buffer.error(f"Sync handler failed: {str(e)}", "HANDLER", {"traceback": tb})
            return {"statusCode": 500, "body": json.dumps({"error": str(e)})}

    def run_callback_handler(self, external_id: str) -> Dict[str, Any]:
        """Run the ARRMS callback handler code."""
        log_buffer.handler(f"Executing callback handler for external_id {external_id}")

        try:
            os.environ["LOCAL_DEV"] = "true"
            os.environ["ONSPRING_API_URL"] = self.mock_onspring_url
            os.environ["ONSPRING_API_KEY_SECRET"] = "onspring-api-key"
            os.environ["ONSPRING_API_KEY"] = "local-test-key"
            os.environ["ARRMS_API_KEY_SECRET"] = "arrms-api-key"

            event = {
                "httpMethod": "POST",
                "path": "/callback/arrms",
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"external_id": external_id, "external_source": "onspring"}),
                "isBase64Encoded": False,
            }

            from handlers.arrms_to_onspring import lambda_handler

            result = lambda_handler(event, self.create_mock_context())

            log_buffer.handler(f"Callback handler completed", {"status_code": result.get("statusCode")})
            return result

        except Exception as e:
            tb = traceback.format_exc()
            log_buffer.error(f"Callback handler failed: {str(e)}", "HANDLER", {"traceback": tb})
            return {"statusCode": 500, "body": json.dumps({"error": str(e)})}


# ============================================================================
# Status Cache
# ============================================================================

arrms_status_cache = {
    "connected": False,
    "error": "Not checked yet",
    "last_check": None,
}


# ============================================================================
# HTTP Server
# ============================================================================


class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    """HTTP server that handles each request in a separate thread."""

    daemon_threads = True


class CommandCenterHandler(BaseHTTPRequestHandler):
    """HTTP request handler for the command center."""

    # Class-level references (set by main)
    arrms_client: Optional[LocalARRMSClient] = None
    mock_client: Optional[MockOnspringClient] = None
    mock_server: Optional[MockOnspringServer] = None
    handler_executor: Optional[HandlerExecutor] = None
    web_ui_path: str = ""

    def log_message(self, format, *args):
        """Suppress default logging."""
        pass

    def send_json(self, data: Any, status: int = 200):
        """Send JSON response."""
        try:
            self.send_response(status)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps(data).encode())
        except BrokenPipeError:
            pass

    def send_html(self, content: str):
        """Send HTML response."""
        try:
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(content.encode())
        except BrokenPipeError:
            pass

    def do_OPTIONS(self):
        """Handle CORS preflight."""
        try:
            self.send_response(200)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "GET, POST, DELETE, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")
            self.end_headers()
        except BrokenPipeError:
            pass

    def do_GET(self):
        """Handle GET requests."""
        try:
            parsed = urlparse(self.path)
            path = parsed.path

            if path == "/" or path == "/index.html":
                self.serve_ui()
            elif path == "/api/logs":
                self.serve_logs()
            elif path == "/api/logs/stream":
                self.serve_sse()
            elif path == "/api/status":
                self.serve_status()
            elif path == "/api/mock-records":
                self.serve_mock_records()
            elif path == "/api/onspring-updates":
                self.serve_onspring_updates()
            elif path == "/docs/arrms":
                self.serve_swagger_ui("ARRMS API", "/docs/arrms/openapi.yaml")
            elif path == "/docs/onspring":
                self.serve_swagger_ui("Onspring API", "/docs/onspring/openapi.yaml")
            elif path == "/docs/arrms/openapi.yaml":
                self.serve_openapi_spec("arrms_openapi.yaml")
            elif path == "/docs/onspring/openapi.yaml":
                self.serve_openapi_spec("onspring_openapi.yaml")
            elif path == "/health":
                self.send_json({"status": "ok"})
            else:
                self.send_response(404)
                self.end_headers()
        except BrokenPipeError:
            pass

    def do_POST(self):
        """Handle POST requests."""
        try:
            parsed = urlparse(self.path)
            path = parsed.path

            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length).decode("utf-8") if content_length > 0 else ""

            try:
                payload = json.loads(body) if body else {}
            except json.JSONDecodeError:
                payload = {}

            if path == "/api/send":
                self.handle_send(payload)
            elif path == "/api/stats":
                self.handle_stats(payload)
            elif path == "/api/apikey":
                self.handle_apikey_update(payload)
            elif path == "/api/apiurl":
                self.handle_apiurl_update(payload)
            elif path == "/api/run-handler":
                self.handle_run_handler(payload)
            elif path == "/webhook/arrms":
                self.handle_webhook(payload)
            else:
                self.send_json({"error": "Not found"}, 404)
        except BrokenPipeError:
            pass

    def do_DELETE(self):
        """Handle DELETE requests."""
        try:
            parsed = urlparse(self.path)
            path = parsed.path

            if path == "/api/logs":
                log_buffer.clear()
                self.send_json({"status": "ok", "message": "Logs cleared"})
            elif path == "/api/onspring-updates":
                if self.mock_server:
                    self.mock_server.clear_updates()
                self.send_json({"status": "ok", "message": "Updates cleared"})
            else:
                self.send_json({"error": "Not found"}, 404)
        except BrokenPipeError:
            pass

    def serve_ui(self):
        """Serve the main HTML UI."""
        html_path = os.path.join(self.web_ui_path, "index.html")
        try:
            with open(html_path, "r") as f:
                self.send_html(f.read())
        except FileNotFoundError:
            self.send_html("<h1>UI not found</h1><p>Missing: " + html_path + "</p>")

    def serve_logs(self):
        """Serve recent logs as JSON."""
        logs = log_buffer.get_recent(100)
        self.send_json({"logs": logs})

    def serve_status(self):
        """Serve current status."""
        try:
            arrms_url = os.environ.get("ARRMS_API_URL", "not configured")
            mock_client = CommandCenterHandler.mock_client
            mock_server = CommandCenterHandler.mock_server
            arrms_client = CommandCenterHandler.arrms_client

            # Get masked API key
            api_key_masked = None
            if arrms_client and arrms_client.api_key:
                key = arrms_client.api_key
                api_key_masked = "****" + key[-4:] if len(key) > 4 else "****"

            status = {
                "arrms_connected": arrms_status_cache["connected"],
                "arrms_url": arrms_url,
                "arrms_error": arrms_status_cache["error"],
                "arrms_api_key_masked": api_key_masked,
                "mock_onspring_running": mock_server.running if mock_server else False,
                "mock_onspring_port": mock_server.port if mock_server else None,
                "mock_records_count": len(mock_client.records) if mock_client else 0,
                "onspring_updates_count": len(mock_server.field_updates) if mock_server else 0,
            }
            self.send_json(status)
        except Exception as e:
            self.send_json({"arrms_connected": False, "arrms_error": str(e)})

    def serve_mock_records(self):
        """Serve mock Onspring records."""
        # MOCK_RECORDS can be a list or dict depending on source
        if isinstance(MOCK_RECORDS, dict):
            records = list(MOCK_RECORDS.values())
        else:
            records = list(MOCK_RECORDS)
        self.send_json({"records": records})

    def serve_onspring_updates(self):
        """Serve Onspring field updates."""
        updates = self.mock_server.get_updates() if self.mock_server else []
        self.send_json({"updates": updates})

    def serve_swagger_ui(self, title: str, spec_url: str):
        """Serve Swagger UI for an API."""
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{title} - Swagger UI</title>
    <link rel="stylesheet" href="https://unpkg.com/swagger-ui-dist@5.9.0/swagger-ui.css">
    <style>
        body {{ margin: 0; padding: 0; }}
        .topbar {{ display: none !important; }}
        .swagger-ui .info {{ margin: 20px 0; }}
        .back-link {{
            display: block;
            padding: 10px 20px;
            background: #1a1a2e;
            color: #6366f1;
            text-decoration: none;
            font-family: sans-serif;
        }}
        .back-link:hover {{ color: #818cf8; }}
    </style>
</head>
<body>
    <a class="back-link" href="/">&larr; Back to Command Center</a>
    <div id="swagger-ui"></div>
    <script src="https://unpkg.com/swagger-ui-dist@5.9.0/swagger-ui-bundle.js"></script>
    <script>
        window.onload = function() {{
            SwaggerUIBundle({{
                url: "{spec_url}",
                dom_id: '#swagger-ui',
                presets: [SwaggerUIBundle.presets.apis, SwaggerUIBundle.SwaggerUIStandalonePreset],
                layout: "BaseLayout"
            }});
        }};
    </script>
</body>
</html>"""
        self.send_html(html)

    def serve_openapi_spec(self, filename: str):
        """Serve an OpenAPI spec file."""
        docs_path = os.path.join(os.path.dirname(__file__), "docs", filename)
        try:
            with open(docs_path, "r") as f:
                content = f.read()
            self.send_response(200)
            self.send_header("Content-Type", "application/x-yaml")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(content.encode())
        except FileNotFoundError:
            self.send_json({"error": f"Spec file not found: {filename}"}, 404)

    def serve_sse(self):
        """Serve Server-Sent Events stream for real-time logs."""
        self.send_response(200)
        self.send_header("Content-Type", "text/event-stream")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Connection", "keep-alive")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()

        subscriber = log_buffer.subscribe()

        try:
            self.wfile.write(b"event: connected\ndata: {}\n\n")
            self.wfile.flush()

            while True:
                try:
                    entry = subscriber.get(timeout=30)
                    data = json.dumps(entry)
                    self.wfile.write(f"event: log\ndata: {data}\n\n".encode())
                    self.wfile.flush()
                except queue.Empty:
                    self.wfile.write(b": keepalive\n\n")
                    self.wfile.flush()
        except (BrokenPipeError, ConnectionResetError):
            pass
        finally:
            log_buffer.unsubscribe(subscriber)

    def handle_send(self, payload: Dict):
        """Handle send questionnaire request (direct ARRMS client call)."""
        record_id = payload.get("record_id", 12345)
        app_id = payload.get("app_id", 100)

        log_buffer.info(f"Starting send flow for record {record_id}")

        arrms_client = self.arrms_client
        mock_client = self.mock_client

        def send_task():
            try:
                self._run_send_flow(record_id, app_id, arrms_client, mock_client)
            except Exception as e:
                tb = traceback.format_exc()
                log_buffer.error(f"Send flow failed: {str(e)}", "SYSTEM", {"traceback": tb})

        thread = threading.Thread(target=send_task, daemon=True)
        thread.start()

        self.send_json({"status": "started", "record_id": record_id})

    def _run_send_flow(self, record_id: int, app_id: int, arrms_client, mock_client):
        """Execute the send questionnaire flow."""
        log_buffer.info(f"Fetching record {record_id} from mock Onspring")

        record_data = mock_client.get_record(app_id=app_id, record_id=record_id)
        log_buffer.info(f"Got record: {record_data.get('fields', {}).get('Title', {}).get('value', 'Unknown')}")

        transformed = transform_record(record_data, onspring_client=mock_client)
        log_buffer.debug("Transformed record data", "SYSTEM", transformed)

        files = mock_client.get_record_files(record_data)
        if not files:
            log_buffer.error("No files found in record")
            return

        log_buffer.info(f"Found {len(files)} files in record")

        questionnaire_file = files[0]
        file_content = mock_client.download_file(
            record_id=questionnaire_file["record_id"],
            field_id=questionnaire_file["field_id"],
            file_id=questionnaire_file["file_id"],
        )

        file_name = questionnaire_file.get("file_name", "questionnaire.xlsx")
        _, file_ext = os.path.splitext(file_name)
        if not file_ext:
            file_ext = ".xlsx"

        with tempfile.NamedTemporaryFile(mode="wb", suffix=file_ext, delete=False) as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name

        log_buffer.info(f"Saved questionnaire to temp file ({len(file_content)} bytes)")

        try:
            log_buffer.arrms("Checking for existing questionnaire...")
            existing = arrms_client.find_questionnaire_by_external_id(
                external_id=str(record_id),
                external_source="onspring",
            )

            if existing:
                arrms_id = existing.get("id")
                log_buffer.arrms(f"Found existing questionnaire {arrms_id}, updating...")
                result = arrms_client.update_questionnaire_file(
                    questionnaire_id=arrms_id,
                    file_path=temp_file_path,
                    external_metadata=transformed.get("external_metadata", {}),
                )
                log_buffer.arrms(f"Updated questionnaire {arrms_id}", {"result": result})
            else:
                log_buffer.arrms("No existing questionnaire, creating new...")
                result = arrms_client.upload_questionnaire(
                    file_path=temp_file_path,
                    external_id=str(record_id),
                    external_source="onspring",
                    external_metadata=transformed.get("external_metadata", {}),
                )
                arrms_id = result.get("id")
                log_buffer.arrms(f"Created questionnaire {arrms_id}", {"result": result})

            log_buffer.info(f"Send flow completed successfully for record {record_id}")

        finally:
            os.unlink(temp_file_path)

    def handle_stats(self, payload: Dict):
        """Handle get statistics request."""
        record_id = payload.get("record_id", 12345)

        log_buffer.arrms(f"Fetching statistics for record {record_id}")

        arrms_client = self.arrms_client

        def stats_task():
            try:
                stats = arrms_client.get_questionnaire_statistics(
                    external_id=str(record_id),
                    external_source="onspring",
                )
                summary = stats.get("summary", {})
                log_buffer.arrms(
                    f"Stats: {summary.get('approved_questions', 0)}/{summary.get('total_questions', 0)} approved",
                    stats,
                )
            except Exception as e:
                tb = traceback.format_exc()
                log_buffer.error(f"Failed to get statistics: {str(e)}", "ARRMS", {"traceback": tb})

        thread = threading.Thread(target=stats_task, daemon=True)
        thread.start()

        self.send_json({"status": "started", "record_id": record_id})

    def handle_apikey_update(self, payload: Dict):
        """Handle API key update request."""
        new_key = payload.get("api_key", "").strip()

        if not new_key:
            self.send_json({"status": "error", "error": "No API key provided"}, 400)
            return

        arrms_client = CommandCenterHandler.arrms_client
        if not arrms_client:
            self.send_json({"status": "error", "error": "ARRMS client not initialized"}, 500)
            return

        old_key_masked = "****" + arrms_client.api_key[-4:] if len(arrms_client.api_key) > 4 else "****"
        new_key_masked = "****" + new_key[-4:] if len(new_key) > 4 else "****"

        arrms_client.api_key = new_key
        arrms_client.session.headers.update({"X-API-Key": new_key})

        log_buffer.info(f"API key updated: {old_key_masked} -> {new_key_masked}")

        arrms_status_cache["error"] = "Checking..."
        arrms_status_cache["connected"] = False

        self.send_json({
            "status": "ok",
            "message": f"API key updated to {new_key_masked}",
            "api_key_masked": new_key_masked,
        })

    def handle_apiurl_update(self, payload: Dict):
        """Handle API URL update request."""
        new_url = payload.get("api_url", "").strip()

        if not new_url:
            self.send_json({"status": "error", "error": "No API URL provided"}, 400)
            return

        # Basic URL validation
        if not new_url.startswith(("http://", "https://")):
            self.send_json({"status": "error", "error": "URL must start with http:// or https://"}, 400)
            return

        arrms_client = CommandCenterHandler.arrms_client
        if not arrms_client:
            self.send_json({"status": "error", "error": "ARRMS client not initialized"}, 500)
            return

        old_url = arrms_client.base_url
        arrms_client.base_url = new_url.rstrip("/")

        # Also update environment variable for handlers
        os.environ["ARRMS_API_URL"] = arrms_client.base_url

        log_buffer.info(f"ARRMS URL updated: {old_url} -> {arrms_client.base_url}")

        arrms_status_cache["error"] = "Checking..."
        arrms_status_cache["connected"] = False

        self.send_json({
            "status": "ok",
            "message": f"API URL updated to {arrms_client.base_url}",
            "api_url": arrms_client.base_url,
        })

    def handle_run_handler(self, payload: Dict):
        """Handle run actual handler request."""
        handler_type = payload.get("handler", "webhook")
        record_id = payload.get("record_id", 12345)
        app_id = payload.get("app_id", 248)

        executor = self.handler_executor
        if not executor:
            self.send_json({"status": "error", "error": "Handler executor not initialized"}, 500)
            return

        def run_task():
            try:
                if handler_type == "webhook":
                    result = executor.run_webhook_handler(record_id, app_id)
                elif handler_type == "sync":
                    result = executor.run_sync_handler(app_id)
                elif handler_type == "callback":
                    result = executor.run_callback_handler(str(record_id))
                else:
                    log_buffer.error(f"Unknown handler type: {handler_type}")
                    return

                log_buffer.handler(f"Handler result", {"result": result})

            except Exception as e:
                tb = traceback.format_exc()
                log_buffer.error(f"Handler execution failed: {str(e)}", "HANDLER", {"traceback": tb})

        thread = threading.Thread(target=run_task, daemon=True)
        thread.start()

        self.send_json({"status": "started", "handler": handler_type, "record_id": record_id})

    def handle_webhook(self, payload: Dict):
        """Handle incoming ARRMS webhook."""
        event_type = payload.get("event_type", "unknown")
        questionnaire_id = payload.get("questionnaire_id")
        external_refs = payload.get("external_references", [])

        log_buffer.webhook(
            f"Received: {event_type}",
            {"questionnaire_id": questionnaire_id, "external_references": external_refs, "full_payload": payload},
        )

        onspring_ref = None
        for ref in external_refs:
            if ref.get("external_source") == "onspring":
                onspring_ref = ref
                break

        if onspring_ref:
            external_id = onspring_ref.get("external_id")
            log_buffer.webhook(f"Webhook for Onspring record: {external_id}")

            try:
                stats = CommandCenterHandler.arrms_client.get_questionnaire_statistics(
                    external_id=external_id,
                    external_source="onspring",
                )
                summary = stats.get("summary", {})
                log_buffer.arrms(
                    f"ARRMS Status: {summary.get('approved_questions', 0)}/{summary.get('total_questions', 0)} approved",
                    {"summary": summary},
                )
            except Exception as e:
                log_buffer.warning(f"Could not fetch statistics: {str(e)}")

        self.send_json({"status": "received", "event_type": event_type})


# ============================================================================
# Background Tasks
# ============================================================================


def start_arrms_health_checker(arrms_client, interval: int = 5):
    """Start background thread that checks ARRMS health periodically."""

    def check_health():
        while True:
            try:
                response = arrms_client.session.get(f"{arrms_client.base_url}/health", timeout=3)
                if response.status_code in (200, 404):
                    arrms_status_cache["connected"] = True
                    arrms_status_cache["error"] = None
                else:
                    arrms_status_cache["connected"] = False
                    arrms_status_cache["error"] = f"HTTP {response.status_code}"
            except Exception as e:
                arrms_status_cache["connected"] = False
                arrms_status_cache["error"] = str(e)[:80]

            arrms_status_cache["last_check"] = datetime.now().isoformat()
            time.sleep(interval)

    thread = threading.Thread(target=check_health, daemon=True)
    thread.start()
    return thread


# ============================================================================
# Main
# ============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="ARRMS Integration - Local Development Command Center",
    )
    parser.add_argument("--port", type=int, default=8080, help="Web UI port (default: 8080)")
    parser.add_argument("--mock-port", type=int, default=5001, help="Mock Onspring port (default: 5001)")
    parser.add_argument("--env-file", default=".env.local", help="Environment file (default: .env.local)")
    parser.add_argument("--no-log-bodies", action="store_true", help="Don't log HTTP bodies")
    parser.add_argument("--no-mock-server", action="store_true", help="Don't start embedded mock Onspring server")

    args = parser.parse_args()

    # Load environment
    load_env_file(args.env_file)

    # Set local dev mode
    os.environ["LOCAL_DEV"] = "true"
    os.environ["POWERTOOLS_SERVICE_NAME"] = "arrms-integration-local"

    # Validate
    if not os.environ.get("ARRMS_API_URL"):
        logger.error("ARRMS_API_URL not set")
        sys.exit(1)
    if not os.environ.get("ARRMS_API_KEY"):
        logger.error("ARRMS_API_KEY not set")
        sys.exit(1)

    # Initialize mock Onspring server
    mock_server = None
    if not args.no_mock_server:
        mock_server = MockOnspringServer(port=args.mock_port)
        mock_server.start()
        os.environ["ONSPRING_API_URL"] = f"http://127.0.0.1:{args.mock_port}"

    # Initialize clients
    mock_client = MockOnspringClient()
    arrms_client = LocalARRMSClient(log_bodies=not args.no_log_bodies)

    # Initialize handler executor
    mock_onspring_url = f"http://127.0.0.1:{args.mock_port}" if mock_server else os.environ.get("ONSPRING_API_URL", "")
    handler_executor = HandlerExecutor(mock_onspring_url, arrms_client)

    # Set up handler
    CommandCenterHandler.arrms_client = arrms_client
    CommandCenterHandler.mock_client = mock_client
    CommandCenterHandler.mock_server = mock_server
    CommandCenterHandler.handler_executor = handler_executor
    CommandCenterHandler.web_ui_path = os.path.join(os.path.dirname(__file__), "web_ui")

    # Start background ARRMS health checker
    start_arrms_health_checker(arrms_client, interval=5)

    # Start server
    server = ThreadingHTTPServer(("", args.port), CommandCenterHandler)

    log_buffer.info(f"Command Center started on http://localhost:{args.port}")
    log_buffer.info(f"ARRMS URL: {os.environ.get('ARRMS_API_URL')}")
    if mock_server:
        log_buffer.info(f"Mock Onspring: http://127.0.0.1:{args.mock_port}")
    log_buffer.info(f"Webhook endpoint: http://localhost:{args.port}/webhook/arrms")

    print(f"\n{'=' * 60}")
    print(f"  ARRMS Integration - Local Development Command Center")
    print(f"{'=' * 60}")
    print(f"  Web UI:         http://localhost:{args.port}")
    if mock_server:
        print(f"  Mock Onspring:  http://127.0.0.1:{args.mock_port}")
    print(f"  Webhook:        http://localhost:{args.port}/webhook/arrms")
    print(f"  Health:         http://localhost:{args.port}/health")
    print(f"{'=' * 60}")
    print(f"  Press Ctrl+C to stop")
    print(f"{'=' * 60}\n")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down...")
        if mock_server:
            mock_server.stop()
        server.shutdown()


if __name__ == "__main__":
    main()
