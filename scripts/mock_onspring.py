"""
Mock Onspring Client for Local Development

Provides mock Onspring API responses for testing the ARRMS integration
without requiring actual Onspring API access.
"""

import os
from datetime import datetime
from typing import Any, Dict, List, Optional


# Sample mock records that simulate Onspring data
# Mock reference data for External Requestor Company Name (App 249)
MOCK_COMPANIES = {
    501: {"name": "Acme Corporation", "field_14949": "Acme Corporation"},
    502: {"name": "Healthcare Inc", "field_14949": "Healthcare Inc"},
    503: {"name": "TechStart LLC", "field_14949": "TechStart LLC"},
}

MOCK_RECORDS = [
    {
        "recordId": 12345,
        "appId": 100,
        # Legacy format (kept for backward compatibility)
        "fields": {
            "Title": {"value": "SOC 2 Type II Assessment", "fieldId": 101},
            "Client": {"value": "Acme Corporation", "fieldId": 102},
            "DueDate": {"value": "2026-03-31", "fieldId": 103},
            "Status": {"value": "New", "fieldId": 104},
            "Description": {"value": "Annual SOC 2 Type II assessment for cloud services", "fieldId": 105},
        },
        # New format matching production Onspring API
        "fieldData": [
            {"type": "String", "fieldId": 101, "value": "SOC 2 Type II Assessment"},
            {"type": "String", "fieldId": 102, "value": "Acme Corporation"},
            {"type": "Date", "fieldId": 103, "value": "2026-03-31"},
            {"type": "String", "fieldId": 104, "value": "New"},
            {"type": "String", "fieldId": 105, "value": "Annual SOC 2 Type II assessment for cloud services"},
            # New fields matching production field IDs
            {"type": "Date", "fieldId": 14872, "value": "2026-03-31"},  # Request Due Back to External Requestor
            {"type": "String", "fieldId": 14888, "value": "SOC 2 Type II scope: Cloud infrastructure and security controls"},  # Scope Summary
            {"type": "Integer", "fieldId": 14947, "value": 501},  # External Requestor Company Name (reference to app 249)
            {"type": "String", "fieldId": 15083, "value": None},  # Questionnaire Link (written back by integration)
            {
                "type": "AttachmentList",
                "fieldId": 200,
                "value": [
                    {
                        "fileId": 1001,
                        "fileName": "questionnaire.xlsx",
                        "fileSize": 15000,
                        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "notes": "Main questionnaire file",
                    },
                    {
                        "fileId": 1002,
                        "fileName": "supporting_docs.pdf",
                        "fileSize": 25000,
                        "contentType": "application/pdf",
                        "notes": "Supporting documentation",
                    },
                ],
            },
        ],
    },
    {
        "recordId": 12346,
        "appId": 100,
        # Legacy format (kept for backward compatibility)
        "fields": {
            "Title": {"value": "HIPAA Security Assessment", "fieldId": 101},
            "Client": {"value": "Healthcare Inc", "fieldId": 102},
            "DueDate": {"value": "2026-04-15", "fieldId": 103},
            "Status": {"value": "In Progress", "fieldId": 104},
            "Description": {"value": "HIPAA security risk assessment", "fieldId": 105},
        },
        # New format matching production Onspring API
        "fieldData": [
            {"type": "String", "fieldId": 101, "value": "HIPAA Security Assessment"},
            {"type": "String", "fieldId": 102, "value": "Healthcare Inc"},
            {"type": "Date", "fieldId": 103, "value": "2026-04-15"},
            {"type": "String", "fieldId": 104, "value": "In Progress"},
            {"type": "String", "fieldId": 105, "value": "HIPAA security risk assessment"},
            # New fields matching production field IDs
            {"type": "Date", "fieldId": 14872, "value": "2026-04-15"},  # Request Due Back to External Requestor
            {"type": "String", "fieldId": 14888, "value": "HIPAA security risk assessment for patient data systems"},  # Scope Summary
            {"type": "Integer", "fieldId": 14947, "value": 502},  # External Requestor Company Name (reference to app 249)
            {"type": "String", "fieldId": 15083, "value": None},  # Questionnaire Link (written back by integration)
            {
                "type": "AttachmentList",
                "fieldId": 200,
                "value": [
                    {
                        "fileId": 2001,
                        "fileName": "hipaa_questionnaire.xlsx",
                        "fileSize": 18000,
                        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "notes": "HIPAA questionnaire",
                    },
                ],
            },
        ],
    },
]


class MockOnspringClient:
    """
    Mock Onspring client that returns predefined test data.

    Simulates the OnspringClient interface without making actual API calls.
    """

    def __init__(self, mock_records: Optional[List[Dict[str, Any]]] = None):
        """
        Initialize mock client with optional custom records.

        Args:
            mock_records: Optional list of mock records. Uses MOCK_RECORDS if not provided.
        """
        self.records = mock_records or MOCK_RECORDS
        self._file_contents: Dict[int, bytes] = {}
        self._load_sample_files()
        print("[MockOnspringClient] Initialized with {} mock records".format(len(self.records)))

    def _load_sample_files(self):
        """Load sample files from scripts/sample_files directory."""
        sample_dir = os.path.join(os.path.dirname(__file__), "sample_files")

        if os.path.exists(sample_dir):
            for filename in os.listdir(sample_dir):
                filepath = os.path.join(sample_dir, filename)
                if os.path.isfile(filepath):
                    with open(filepath, "rb") as f:
                        # Use hash of filename as mock file ID
                        file_id = abs(hash(filename)) % 10000 + 1000
                        self._file_contents[file_id] = f.read()
                        print(f"[MockOnspringClient] Loaded sample file: {filename} (mock ID: {file_id})")

    def health_check(self) -> bool:
        """Mock health check - always returns True."""
        print("[MockOnspringClient] Health check: OK")
        return True

    def get_record(self, app_id: int, record_id: int) -> Dict[str, Any]:
        """
        Get a single mock record by ID.

        Args:
            app_id: Onspring application ID
            record_id: Record ID to retrieve

        Returns:
            Mock record data
        """
        print(f"[MockOnspringClient] Getting record {record_id} from app {app_id}")

        for record in self.records:
            if record["recordId"] == record_id:
                return record

        # Return first record if specific ID not found
        print(f"[MockOnspringClient] Record {record_id} not found, returning first mock record")
        return self.records[0] if self.records else {}

    def get_records(
        self,
        app_id: int,
        filter_criteria: Optional[Dict[str, Any]] = None,
        page_size: int = 100,
        page_number: int = 1,
    ) -> List[Dict[str, Any]]:
        """
        Get multiple mock records.

        Args:
            app_id: Onspring application ID
            filter_criteria: Optional filter (ignored in mock)
            page_size: Number of records per page
            page_number: Page number

        Returns:
            List of mock records
        """
        print(f"[MockOnspringClient] Querying records from app {app_id}")
        print(f"[MockOnspringClient] Filter: {filter_criteria}, page_size: {page_size}")

        # Simple pagination
        start = (page_number - 1) * page_size
        end = start + page_size
        records = self.records[start:end]

        print(f"[MockOnspringClient] Returning {len(records)} records")
        return records

    def get_record_files(self, record_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Extract file attachment information from record.

        Args:
            record_data: Record data dictionary

        Returns:
            List of file attachment metadata
        """
        files = []
        field_data = record_data.get("fieldData", [])
        record_id = record_data.get("recordId")

        for field in field_data:
            value = field.get("value")
            field_id = field.get("fieldId")

            if isinstance(value, list) and len(value) > 0:
                first_item = value[0]
                if isinstance(first_item, dict) and "fileId" in first_item:
                    for file_item in value:
                        file_info = {
                            "record_id": record_id,
                            "field_id": field_id,
                            "file_id": file_item.get("fileId"),
                            "file_name": file_item.get("fileName"),
                            "file_size": file_item.get("fileSize"),
                            "content_type": file_item.get("contentType"),
                            "notes": file_item.get("notes"),
                        }
                        files.append(file_info)

        print(f"[MockOnspringClient] Found {len(files)} files in record {record_id}")
        return files

    def download_file(self, record_id: int, field_id: int, file_id: int) -> bytes:
        """
        Download file content.

        Returns sample file content or generates mock content.

        Args:
            record_id: Record ID containing the file
            field_id: Field ID of the file field
            file_id: File ID to download

        Returns:
            File content as bytes
        """
        print(f"[MockOnspringClient] Downloading file {file_id} from record {record_id}")

        # Check if we have loaded sample file content
        if self._file_contents:
            # Return first available sample file
            content = next(iter(self._file_contents.values()))
            print(f"[MockOnspringClient] Returning sample file content ({len(content)} bytes)")
            return content

        # Generate mock Excel-like content if no sample files
        mock_content = self._generate_mock_excel()
        print(f"[MockOnspringClient] Generated mock file content ({len(mock_content)} bytes)")
        return mock_content

    def _generate_mock_excel(self) -> bytes:
        """Generate minimal mock Excel file content."""
        try:
            from openpyxl import Workbook
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "Questionnaire"

            # Add some sample data
            ws["A1"] = "Question"
            ws["B1"] = "Response"
            ws["C1"] = "Notes"
            ws["A2"] = "What is your company name?"
            ws["A3"] = "Describe your security policies"
            ws["A4"] = "Do you have SOC 2 certification?"

            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        except ImportError:
            # Fallback: return minimal valid xlsx bytes (empty workbook)
            # This is a minimal valid XLSX file structure
            print("[MockOnspringClient] openpyxl not installed, using minimal mock content")
            return b"PK\x03\x04" + b"\x00" * 100  # Minimal zip header

    def get_file_info(self, record_id: int, field_id: int, file_id: int) -> Dict[str, Any]:
        """
        Get file metadata.

        Args:
            record_id: Record ID containing the file
            field_id: Field ID of the file field
            file_id: File ID

        Returns:
            File metadata dictionary
        """
        return {
            "fileId": file_id,
            "fileName": f"mock_file_{file_id}.xlsx",
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "fileSize": 15000,
        }

    def resolve_reference_field(
        self, referenced_app_id: int, referenced_record_id: int, field_id: int
    ) -> Optional[str]:
        """
        Resolve a reference field value from another app.

        Used to get the company name from the External Requestor Company Name field.

        Args:
            referenced_app_id: App ID of the referenced record (e.g., 249 for companies)
            referenced_record_id: Record ID in the referenced app
            field_id: Field ID to retrieve from the referenced record

        Returns:
            Field value as string, or None if not found
        """
        print(f"[MockOnspringClient] Resolving reference: app={referenced_app_id}, record={referenced_record_id}, field={field_id}")

        # Look up in mock companies data
        if referenced_app_id == 249 and referenced_record_id in MOCK_COMPANIES:
            company = MOCK_COMPANIES[referenced_record_id]
            value = company.get(f"field_{field_id}") or company.get("name")
            print(f"[MockOnspringClient] Resolved company name: {value}")
            return value

        print(f"[MockOnspringClient] Reference not found, returning None")
        return None

    def update_field_value(
        self, app_id: int, record_id: int, field_id: int, value: Any
    ) -> Dict[str, Any]:
        """
        Update a single field value on a record.

        Args:
            app_id: Application ID
            record_id: Record ID to update
            field_id: Field ID to update
            value: New value for the field

        Returns:
            Update response
        """
        print(f"[MockOnspringClient] Updating field {field_id} on record {record_id} in app {app_id}")
        print(f"[MockOnspringClient] New value: {value}")

        # Find and update the record in mock data (for visibility in local testing)
        for record in self.records:
            if record["recordId"] == record_id:
                for field in record.get("fieldData", []):
                    if field.get("fieldId") == field_id:
                        field["value"] = value
                        print(f"[MockOnspringClient] Updated field {field_id} in mock record")
                        break

        return {"id": record_id, "warnings": []}


def create_mock_webhook_payload(record_id: int = 12345, app_id: int = 100) -> Dict[str, Any]:
    """
    Create a mock webhook payload matching Onspring's format.

    Args:
        record_id: Record ID for the webhook
        app_id: Application ID

    Returns:
        API Gateway-style event with webhook body
    """
    import json

    return {
        "resource": "/webhook/onspring",
        "path": "/webhook/onspring",
        "httpMethod": "POST",
        "headers": {
            "Content-Type": "application/json",
            "x-api-key": "local-test-key",
        },
        "body": json.dumps([{"RecordId": str(record_id), "AppId": str(app_id)}]),
        "isBase64Encoded": False,
    }


def create_mock_sync_payload(
    app_id: int = 100,
    filter_criteria: Optional[Dict[str, Any]] = None,
    batch_size: int = 50,
) -> Dict[str, Any]:
    """
    Create a mock sync request payload.

    Args:
        app_id: Application ID
        filter_criteria: Optional filter
        batch_size: Batch size

    Returns:
        API Gateway-style event with sync request body
    """
    import json

    body = {
        "app_id": app_id,
        "batch_size": batch_size,
    }
    if filter_criteria:
        body["filter"] = filter_criteria

    return {
        "resource": "/sync/onspring-to-arrms",
        "path": "/sync/onspring-to-arrms",
        "httpMethod": "POST",
        "headers": {
            "Content-Type": "application/json",
            "x-api-key": "local-test-key",
        },
        "body": json.dumps(body),
        "isBase64Encoded": False,
    }
