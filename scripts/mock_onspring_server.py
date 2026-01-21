#!/usr/bin/env python3
"""
Mock Onspring HTTP Server

A Flask-based mock server that implements Onspring API endpoints for local development.
Run this server to test the integration without requiring actual Onspring access.

Usage:
    python scripts/mock_onspring_server.py

    # Or with custom port
    MOCK_ONSPRING_PORT=5001 python scripts/mock_onspring_server.py

The server will start on http://localhost:5001 by default.
Set ONSPRING_API_URL=http://localhost:5001 to use with the integration.
"""

import io
import json
import os
import sys
from datetime import datetime
from typing import Any, Dict, List, Optional

from flask import Flask, jsonify, request, send_file

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

app = Flask(__name__)

# ============================================================================
# Mock Data Store
# ============================================================================

# Mock companies for reference field resolution (App 249)
MOCK_COMPANIES = {
    501: {
        "recordId": 501,
        "appId": 249,
        "fieldData": [
            {"fieldId": 14949, "type": "String", "value": "Acme Corporation"},
        ],
    },
    502: {
        "recordId": 502,
        "appId": 249,
        "fieldData": [
            {"fieldId": 14949, "type": "String", "value": "Healthcare Inc"},
        ],
    },
    503: {
        "recordId": 503,
        "appId": 249,
        "fieldData": [
            {"fieldId": 14949, "type": "String", "value": "TechStart LLC"},
        ],
    },
}

# Mock questionnaire records (App 248 or configurable)
MOCK_RECORDS: Dict[int, Dict[str, Any]] = {
    12345: {
        "recordId": 12345,
        "appId": 248,
        "fieldData": [
            {"fieldId": 101, "type": "String", "value": "SOC 2 Type II Assessment"},
            {"fieldId": 102, "type": "String", "value": "Acme Corporation"},
            {"fieldId": 103, "type": "Date", "value": "2026-03-31"},
            {"fieldId": 104, "type": "String", "value": "New"},
            {"fieldId": 105, "type": "String", "value": "Annual SOC 2 Type II assessment for cloud services"},
            {"fieldId": 14872, "type": "Date", "value": "2026-03-31"},
            {"fieldId": 14888, "type": "String", "value": "SOC 2 Type II scope: Cloud infrastructure and security controls"},
            {"fieldId": 14947, "type": "Integer", "value": 501},
            {"fieldId": 15083, "type": "String", "value": None},
            {
                "fieldId": 200,
                "type": "AttachmentList",
                "value": [
                    {
                        "fileId": 1001,
                        "fileName": "questionnaire.xlsx",
                        "fileSize": 15000,
                        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "notes": "Main questionnaire file",
                    },
                ],
            },
        ],
    },
    12346: {
        "recordId": 12346,
        "appId": 248,
        "fieldData": [
            {"fieldId": 101, "type": "String", "value": "HIPAA Security Assessment"},
            {"fieldId": 102, "type": "String", "value": "Healthcare Inc"},
            {"fieldId": 103, "type": "Date", "value": "2026-04-15"},
            {"fieldId": 104, "type": "String", "value": "In Progress"},
            {"fieldId": 105, "type": "String", "value": "HIPAA security risk assessment"},
            {"fieldId": 14872, "type": "Date", "value": "2026-04-15"},
            {"fieldId": 14888, "type": "String", "value": "HIPAA security assessment for patient data systems"},
            {"fieldId": 14947, "type": "Integer", "value": 502},
            {"fieldId": 15083, "type": "String", "value": None},
            {
                "fieldId": 200,
                "type": "AttachmentList",
                "value": [
                    {
                        "fileId": 2001,
                        "fileName": "hipaa_questionnaire.xlsx",
                        "fileSize": 18000,
                        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "notes": "HIPAA questionnaire",
                    },
                ],
            },
        ],
    },
}

# Track field updates for verification
FIELD_UPDATES: List[Dict[str, Any]] = []


# ============================================================================
# Helper Functions
# ============================================================================


def get_field_value(record: Dict[str, Any], field_id: int) -> Any:
    """Extract field value from record by field ID."""
    for field in record.get("fieldData", []):
        if field.get("fieldId") == field_id:
            return field.get("value")
    return None


def set_field_value(record: Dict[str, Any], field_id: int, value: Any) -> None:
    """Set field value in record by field ID."""
    for field in record.get("fieldData", []):
        if field.get("fieldId") == field_id:
            field["value"] = value
            return
    # Field doesn't exist, add it
    record["fieldData"].append({"fieldId": field_id, "type": "String", "value": value})


def generate_mock_excel() -> bytes:
    """Generate minimal mock Excel file content."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Questionnaire"

        # Add sample questionnaire data
        ws["A1"] = "Question"
        ws["B1"] = "Response"
        ws["C1"] = "Notes"
        ws["A2"] = "What is your company name?"
        ws["A3"] = "Describe your security policies"
        ws["A4"] = "Do you have SOC 2 certification?"
        ws["A5"] = "What encryption standards do you use?"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except ImportError:
        # Fallback: return minimal content
        return b"Mock Excel Content - install openpyxl for real Excel files"


def load_sample_file(filename: str) -> Optional[bytes]:
    """Load sample file from scripts/sample_files directory."""
    sample_dir = os.path.join(os.path.dirname(__file__), "sample_files")
    filepath = os.path.join(sample_dir, filename)

    if os.path.exists(filepath):
        with open(filepath, "rb") as f:
            return f.read()
    return None


# ============================================================================
# API Endpoints
# ============================================================================


@app.route("/Ping", methods=["GET"])
def ping():
    """Health check endpoint."""
    return jsonify({"status": "ok", "message": "Mock Onspring Server"})


@app.route("/Records/appId/<int:app_id>/recordId/<int:record_id>", methods=["GET"])
def get_record(app_id: int, record_id: int):
    """
    Get a single record by ID.

    Onspring API: GET /Records/appId/{appId}/recordId/{recordId}
    """
    print(f"[MockOnspring] GET record {record_id} from app {app_id}")

    # Check companies first (app 249)
    if app_id == 249 and record_id in MOCK_COMPANIES:
        return jsonify(MOCK_COMPANIES[record_id])

    # Check questionnaire records
    if record_id in MOCK_RECORDS:
        record = MOCK_RECORDS[record_id]
        # Ensure app_id matches or return anyway for flexibility
        return jsonify(record)

    # Return first record if not found (for testing flexibility)
    if MOCK_RECORDS:
        first_record = list(MOCK_RECORDS.values())[0]
        print(f"[MockOnspring] Record {record_id} not found, returning first mock record")
        return jsonify(first_record)

    return jsonify({"error": "Record not found"}), 404


@app.route("/Records", methods=["PUT"])
def update_record():
    """
    Update a record.

    Onspring API: PUT /Records
    Body: { "appId": 248, "recordId": 12345, "fields": { "fieldId": value } }
    """
    data = request.get_json()
    app_id = data.get("appId")
    record_id = data.get("recordId")
    fields = data.get("fields", {})

    print(f"[MockOnspring] PUT record {record_id} in app {app_id}")
    print(f"[MockOnspring] Fields to update: {json.dumps(fields, indent=2)}")

    # Track the update
    update_record = {
        "timestamp": datetime.utcnow().isoformat(),
        "app_id": app_id,
        "record_id": record_id,
        "fields": fields,
    }
    FIELD_UPDATES.append(update_record)

    # Actually update the mock record
    if record_id in MOCK_RECORDS:
        for field_id_str, value in fields.items():
            field_id = int(field_id_str)
            set_field_value(MOCK_RECORDS[record_id], field_id, value)
            print(f"[MockOnspring] Updated field {field_id} = {value}")

    return jsonify({"id": record_id, "warnings": []})


@app.route("/Records/appId/<int:app_id>/recordId/<int:record_id>/fieldId/<int:field_id>", methods=["PUT"])
def update_field(app_id: int, record_id: int, field_id: int):
    """
    Update a single field value.

    Onspring API: PUT /Records/appId/{appId}/recordId/{recordId}/fieldId/{fieldId}
    """
    data = request.get_json()
    value = data.get("value") if data else request.data.decode("utf-8")

    print(f"[MockOnspring] PUT field {field_id} on record {record_id}")
    print(f"[MockOnspring] New value: {value}")

    # Track the update
    update_record = {
        "timestamp": datetime.utcnow().isoformat(),
        "app_id": app_id,
        "record_id": record_id,
        "field_id": field_id,
        "value": value,
    }
    FIELD_UPDATES.append(update_record)

    # Update the mock record
    if record_id in MOCK_RECORDS:
        set_field_value(MOCK_RECORDS[record_id], field_id, value)

    return jsonify({"id": record_id, "warnings": []})


@app.route("/Files/recordId/<int:record_id>/fieldId/<int:field_id>/fileId/<int:file_id>/file", methods=["GET"])
def download_file(record_id: int, field_id: int, file_id: int):
    """
    Download a file attachment.

    Onspring API: GET /Files/recordId/{recordId}/fieldId/{fieldId}/fileId/{fileId}/file
    """
    print(f"[MockOnspring] GET file {file_id} from record {record_id}, field {field_id}")

    # Try to find filename from record
    filename = "questionnaire.xlsx"
    if record_id in MOCK_RECORDS:
        for field in MOCK_RECORDS[record_id].get("fieldData", []):
            if field.get("fieldId") == field_id and isinstance(field.get("value"), list):
                for file_info in field["value"]:
                    if file_info.get("fileId") == file_id:
                        filename = file_info.get("fileName", filename)
                        break

    # Try to load actual sample file
    content = load_sample_file(filename)
    if content is None:
        content = generate_mock_excel()

    return send_file(
        io.BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/Files/recordId/<int:record_id>/fieldId/<int:field_id>/fileId/<int:file_id>", methods=["GET"])
def get_file_info(record_id: int, field_id: int, file_id: int):
    """
    Get file metadata.

    Onspring API: GET /Files/recordId/{recordId}/fieldId/{fieldId}/fileId/{fileId}
    """
    print(f"[MockOnspring] GET file info for {file_id}")

    # Try to find file info from record
    if record_id in MOCK_RECORDS:
        for field in MOCK_RECORDS[record_id].get("fieldData", []):
            if field.get("fieldId") == field_id and isinstance(field.get("value"), list):
                for file_info in field["value"]:
                    if file_info.get("fileId") == file_id:
                        return jsonify(file_info)

    # Return generic file info
    return jsonify({
        "fileId": file_id,
        "fileName": f"file_{file_id}.xlsx",
        "fileSize": 15000,
        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    })


# ============================================================================
# Admin/Debug Endpoints
# ============================================================================


@app.route("/admin/records", methods=["GET"])
def list_records():
    """List all mock records (admin endpoint)."""
    return jsonify({"records": list(MOCK_RECORDS.values())})


@app.route("/admin/records/<int:record_id>", methods=["POST"])
def add_or_update_record(record_id: int):
    """Add or update a mock record (admin endpoint)."""
    data = request.get_json()
    data["recordId"] = record_id
    MOCK_RECORDS[record_id] = data
    return jsonify({"status": "ok", "record_id": record_id})


@app.route("/admin/updates", methods=["GET"])
def list_updates():
    """List all field updates that have been made (for verification)."""
    return jsonify({"updates": FIELD_UPDATES})


@app.route("/admin/updates", methods=["DELETE"])
def clear_updates():
    """Clear the update history."""
    FIELD_UPDATES.clear()
    return jsonify({"status": "ok", "message": "Updates cleared"})


@app.route("/admin/reset", methods=["POST"])
def reset_state():
    """Reset all mock data to initial state."""
    FIELD_UPDATES.clear()
    # Reset records to initial state would require storing initial state
    return jsonify({"status": "ok", "message": "State reset"})


# ============================================================================
# Main
# ============================================================================


def main():
    """Run the mock Onspring server."""
    port = int(os.environ.get("MOCK_ONSPRING_PORT", 5001))
    host = os.environ.get("MOCK_ONSPRING_HOST", "127.0.0.1")

    print("=" * 60)
    print("Mock Onspring Server")
    print("=" * 60)
    print(f"Starting on http://{host}:{port}")
    print()
    print("Set this environment variable to use with the integration:")
    print(f"  export ONSPRING_API_URL=http://{host}:{port}")
    print()
    print("Available endpoints:")
    print("  GET  /Ping                                    - Health check")
    print("  GET  /Records/appId/{appId}/recordId/{id}     - Get record")
    print("  PUT  /Records                                 - Update record")
    print("  GET  /Files/.../file                          - Download file")
    print()
    print("Admin endpoints:")
    print("  GET  /admin/records                           - List all records")
    print("  GET  /admin/updates                           - List field updates")
    print("  DELETE /admin/updates                         - Clear update history")
    print("=" * 60)
    print()

    app.run(host=host, port=port, debug=True)


if __name__ == "__main__":
    main()
